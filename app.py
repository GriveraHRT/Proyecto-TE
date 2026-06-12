import streamlit as pd
import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de página de Streamlit
st.set_page_config(
    page_title="Analizador de Tiempos - Laboratorio",
    page_icon="🧪",
    layout="wide"
)

st.title("🧪 Analizador de Tiempos de Operación de Laboratorio")
st.markdown("""
Esta aplicación procesa los archivos de extracción de muestras, calcula con precisión matemática 
el tiempo transcurrido entre el **Ingreso** y la **Extracción** utilizando las marcas de tiempo base, 
y genera un reporte gerencial en formato Excel.
""")

# Componente de carga de archivos
uploaded_file = st.file_uploader("Carga tu archivo de datos (CSV o Excel)", type=["csv", "xlsx", "xls"])

def procesar_datos(file):
    # Lectura dinámica según el tipo de archivo
    if file.name.endswith('.csv'):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
        
    # Verificar columnas críticas
    columnas_requeridas = ['FECHA Y HORA DE INGRESO', 'FECHA Y HORA DE EXTRACCIÓN']
    for col in columnas_requeridas:
        if col not in df.columns:
            st.error(f"El archivo no contiene la columna requerida: '{col}'")
            return None, None, None, None

    # Clonar dataframe para no alterar el original y parsear fechas de forma segura
    df_proc = df.copy()
    df_proc['INGRESO_DT'] = pd.to_datetime(df_proc['FECHA Y HORA DE INGRESO'], dayfirst=True, errors='coerce')
    df_proc['EXTRACCION_DT'] = pd.to_datetime(df_proc['FECHA Y HORA DE EXTRACCIÓN'], dayfirst=True, errors='coerce')
    
    # Calcular el tiempo delta en minutos reales
    df_proc['TIEMPO CALCULADO (MINUTOS)'] = (df_proc['EXTRACCION_DT'] - df_proc['INGRESO_DT']).dt.total_seconds() / 60.0
    
    # Generar dataframes de resumen
    total_muestras = len(df_proc)
    promedio_general = df_proc['TIEMPO CALCULADO (MINUTOS)'].mean()
    
    # Resumen por usuario extractor
    if 'USUARIO QUE EXTRAJO LA MUESTA' in df_proc.columns:
        summary_user = df_proc.groupby('USUARIO QUE EXTRAJO LA MUESTA')['TIEMPO CALCULADO (MINUTOS)'].agg(['count', 'mean']).reset_index()
        summary_user.columns = ['Usuario de Extracción', 'Cantidad de Muestras', 'Tiempo Promedio (Minutos)']
        summary_user = summary_user.sort_values(by='Tiempo Promedio (Minutos)', ascending=False)
    else:
        summary_user = pd.DataFrame()

    # Resumen por procedencia
    if 'PROCEDENCIA' in df_proc.columns:
        summary_proc = df_proc.groupby('PROCEDENCIA')['TIEMPO CALCULADO (MINUTOS)'].agg(['count', 'mean']).reset_index()
        summary_proc.columns = ['Procedencia / Servicio', 'Cantidad', 'Tiempo Promedio (Minutos)']
        summary_proc = summary_proc.sort_values(by='Tiempo Promedio (Minutos)', ascending=False)
    else:
        summary_proc = pd.DataFrame()
        
    # Eliminar columnas temporales antes de exportar la data detallada
    df_detalles = df_proc.drop(columns=['INGRESO_DT', 'EXTRACCION_DT'])
    
    return df_detalles, summary_user, summary_proc, total_muestras, promedio_general

def generar_excel_profesional(df_details, summary_user, summary_proc, total_muestras, promedio_general):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Estilos del reporte
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid") # Slate/Navy
    accent_fill = PatternFill(start_color="4A6572", end_color="4A6572", fill_type="solid") # Muted Steel Blue
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    kpi_fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
    
    font_title = Font(name=font_family, size=15, bold=True, color="2C3E50")
    font_section = Font(name=font_family, size=12, bold=True, color="2C3E50")
    font_regular = Font(name=font_family, size=11)
    font_kpi_val = Font(name=font_family, size=18, bold=True, color="2C3E50")
    font_kpi_lbl = Font(name=font_family, size=9, italic=True, color="7F8C8D")
    
    thin_side = Side(border_style="thin", color="BDC3C7")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # ---- HOJA 1: RESUMEN ----
    ws_resumen = wb.active
    ws_resumen.title = "Resumen y Métricas"
    ws_resumen.views.sheetView[0].showGridLines = True
    
    ws_resumen["B2"] = "REPORTE DE TIEMPOS DE OPERACIÓN - LABORATORIO"
    ws_resumen["B2"].font = font_title
    
    # Tarjeta KPI 1: Total
    ws_resumen.merge_cells("B4:C4")
    ws_resumen["B4"] = "TOTAL MUESTRAS PROCESADAS"
    ws_resumen["B4"].font = font_kpi_lbl
    ws_resumen["B4"].alignment = align_center
    ws_resumen["B4"].fill = kpi_fill
    
    ws_resumen.merge_cells("B5:C5")
    ws_resumen["B5"] = total_muestras
    ws_resumen["B5"].font = font_kpi_val
    ws_resumen["B5"].alignment = align_center
    ws_resumen["B5"].fill = kpi_fill
    ws_resumen["B5"].number_format = "#,##0"
    
    # Tarjeta KPI 2: Promedio
    ws_resumen.merge_cells("E4:F4")
    ws_resumen["E4"] = "TIEMPO PROMEDIO GENERAL"
    ws_resumen["E4"].font = font_kpi_lbl
    ws_resumen["E4"].alignment = align_center
    ws_resumen["E4"].fill = kpi_fill
    
    ws_resumen.merge_cells("E5:F5")
    ws_resumen["E5"] = promedio_general
    ws_resumen["E5"].font = font_kpi_val
    ws_resumen["E5"].alignment = align_center
    ws_resumen["E5"].fill = kpi_fill
    ws_resumen["E5"].number_format = "0.00"
    
    ws_resumen["G5"] = "minutos"
    ws_resumen["G5"].font = Font(name=font_family, size=10, italic=True)
    ws_resumen["G5"].alignment = align_left
    
    for r in range(4, 6):
        for c in [2, 3, 5, 6]:
            ws_resumen.cell(row=r, column=c).border = border_all
            
    # Tabla Usuarios
    ws_resumen["B8"] = "Promedio de Tiempo por Usuario de Extracción"
    ws_resumen["B8"].font = font_section
    
    if not summary_user.empty:
        for c_idx, h in enumerate(summary_user.columns, start=2):
            cell = ws_resumen.cell(row=9, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        r_idx = 10
        for _, row in summary_user.iterrows():
            ws_resumen.cell(row=r_idx, column=2, value=row[0]).alignment = align_left
            ws_resumen.cell(row=r_idx, column=3, value=row[1]).number_format = "#,##0"
            ws_resumen.cell(row=r_idx, column=3).alignment = align_right
            cell_v = ws_resumen.cell(row=r_idx, column=4, value=row[2])
            cell_v.number_format = "0.00"
            cell_v.alignment = align_right
            
            for col_c in range(2, 5):
                c = ws_resumen.cell(row=r_idx, column=col_c)
                c.font = font_regular
                c.border = border_all
                if r_idx % 2 == 1:
                    c.fill = zebra_fill
            r_idx += 1
            
    # Tabla Procedencia
    start_r_proc = r_idx + 2 if 'r_idx' in locals() else 8
    ws_resumen.cell(row=start_r_proc, column=2, value="Promedio de Tiempo por Procedencia / Servicio").font = font_section
    
    if not summary_proc.empty:
        for c_idx, h in enumerate(summary_proc.columns, start=2):
            cell = ws_resumen.cell(row=start_r_proc+1, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = accent_fill
            cell.alignment = align_center
            cell.border = border_all
            
        r_idx = start_r_proc + 2
        for _, row in summary_proc.iterrows():
            ws_resumen.cell(row=r_idx, column=2, value=row[0]).alignment = align_left
            ws_resumen.cell(row=r_idx, column=3, value=row[1]).number_format = "#,##0"
            ws_resumen.cell(row=r_idx, column=3).alignment = align_right
            cell_v = ws_resumen.cell(row=r_idx, column=4, value=row[2])
            cell_v.number_format = "0.00"
            cell_v.alignment = align_right
            
            for col_c in range(2, 5):
                c = ws_resumen.cell(row=r_idx, column=col_c)
                c.font = font_regular
                c.border = border_all
                if r_idx % 2 == 1:
                    c.fill = zebra_fill
            r_idx += 1

    # ---- HOJA 2: DETALLES ----
    ws_datos = wb.create_sheet(title="Datos Detallados")
    ws_datos.views.sheetView[0].showGridLines = True
    ws_datos.freeze_panes = "A2"
    
    for c_idx, col_name in enumerate(df_details.columns, start=1):
        cell = ws_datos.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_all
        
    for r_i, row in df_details.iterrows():
        for c_i, val in enumerate(row, start=1):
            cell = ws_datos.cell(row=r_i+2, column=c_i)
            col_name = df_details.columns[c_i-1]
            
            if col_name == 'TIEMPO CALCULADO (MINUTOS)':
                cell.value = val
                cell.number_format = "0.0"
                cell.alignment = align_right
                cell.fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
            elif isinstance(val, (int, float)) and not pd.isna(val):
                cell.value = val
                cell.alignment = align_right if col_name != 'ORDEN' else align_center
                cell.number_format = "#,##0" if col_name == 'ORDEN' else "0.00"
            else:
                cell.value = "" if pd.isna(val) else str(val)
                cell.alignment = align_center if col_name in ['RUT', 'FECHA Y HORA DE INGRESO', 'FECHA Y HORA DE EXTRACCIÓN'] else align_left
                
            cell.font = font_regular
            cell.border = border_all
            if (r_i+2) % 2 == 1 and col_name != 'TIEMPO CALCULADO (MINUTOS)':
                cell.fill = zebra_fill

    # Ajuste automático del ancho de columnas
    for ws in [ws_resumen, ws_datos]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and not (ws == ws_resumen and cell.row == 2):
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output)
    return output.getvalue()

if uploaded_file is not None:
    with st.spinner("Procesando datos y estructurando matrices lógicas..."):
        df_details, summary_user, summary_proc, total_m, prom_g = procesar_datos(uploaded_file)
        
    if df_details is not None:
        st.success("¡Datos calculados con éxito utilizando las marcas de tiempo base!")
        
        # Mostrar métricas en la interfaz
        m1, m2 = st.columns(2)
        m1.metric("Total Órdenes Analizadas", f"{total_m:,}")
        m2.metric("Tiempo Promedio General", f"{prom_g:.2f} minutos")
        
        # Vistas previas en la app
        st.subheader("📊 Vista Previa de Tiempos por Usuario Extractor")
        st.dataframe(summary_user, use_container_width=True)
        
        # Generar el binario del archivo Excel estructurado
        excel_data = generar_excel_profesional(df_details, summary_user, summary_proc, total_m, prom_g)
        
        # Botón para descargar el Excel completo
        st.download_button(
            label="📥 Descargar Reporte Corporativo Excel (.xlsx)",
            data=excel_data,
            file_name="Reporte_Tiempos_Laboratorio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )